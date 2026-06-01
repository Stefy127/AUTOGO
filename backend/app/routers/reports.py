from calendar import monthrange
from datetime import date, datetime, time, timedelta
from io import BytesIO
import re
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Response, status
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import simpleSplit
from reportlab.pdfgen import canvas
from sqlalchemy.orm import Session, joinedload

from app.auth import get_current_user
from app.database import get_db
from app.models import (
    Incident,
    IncidentStatus,
    Payment,
    PaymentMethod,
    Technician,
    User,
    UserRole,
    Workshop,
)
from app.schemas import (
    OperationalReportAppliedFilters,
    OperationalReportItem,
    OperationalReportRequest,
    OperationalReportResponse,
    OperationalReportSummary,
    VoiceReportParseRequest,
    VoiceReportParseResponse,
)

router = APIRouter(prefix="/reports", tags=["reports"])


def _normalize_text(text: str) -> str:
    normalized = (text or "").strip().lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
    }
    for old, new in replacements.items():
        normalized = normalized.replace(old, new)
    return normalized


def _parse_voice_filters(raw_text: str) -> tuple[OperationalReportRequest, Optional[str], list[str]]:
    text = _normalize_text(raw_text)
    warnings: list[str] = []
    today = date.today()

    status_keywords = [
        ("esperando ofertas", "waiting_offers"),
        ("en progreso", "in_progress"),
        ("pendientes", "pending"),
        ("pendiente", "pending"),
        ("asignadas", "assigned"),
        ("asignado", "assigned"),
        ("aceptadas", "accepted"),
        ("aceptado", "accepted"),
        ("completadas", "completed"),
        ("completado", "completed"),
        ("resueltas", "completed"),
        ("resuelto", "completed"),
        ("canceladas", "cancelled"),
        ("cancelado", "cancelled"),
        ("progreso", "in_progress"),
    ]
    payment_keywords = [
        ("transferencia", "transfer"),
        ("efectivo", "cash"),
        ("tarjeta", "card"),
        ("card", "card"),
        ("qr", "qr"),
    ]
    type_keywords = [
        ("neumatico", "tire"),
        ("pinchazo", "tire"),
        ("bateria", "battery"),
        ("llanta", "tire"),
        ("choque", "accident"),
        ("accidente", "accident"),
        ("aceite", "oil"),
        ("motor", "engine"),
    ]
    months = {
        "enero": 1,
        "febrero": 2,
        "marzo": 3,
        "abril": 4,
        "mayo": 5,
        "junio": 6,
        "julio": 7,
        "agosto": 8,
        "septiembre": 9,
        "octubre": 10,
        "noviembre": 11,
        "diciembre": 12,
    }

    status_value = next((value for key, value in status_keywords if key in text), None)
    payment_value = next((value for key, value in payment_keywords if key in text), None)
    incident_type_value = next((value for key, value in type_keywords if key in text), None)

    action: Optional[str] = "query"
    if "pdf" in text:
        action = "pdf"
    elif "excel" in text:
        action = "excel"

    start_date: Optional[date] = None
    end_date: Optional[date] = None

    if "hoy" in text:
        start_date = today
        end_date = today
    elif "ayer" in text:
        day = today - timedelta(days=1)
        start_date = day
        end_date = day
    elif "este mes" in text or "mes actual" in text:
        start_date = today.replace(day=1)
        end_date = today
    elif "ultimos 7 dias" in text or "ultimos siete dias" in text:
        start_date = today - timedelta(days=7)
        end_date = today
    elif "ultimos 30 dias" in text or "ultimos treinta dias" in text:
        start_date = today - timedelta(days=30)
        end_date = today
    else:
        for month_name, month_number in months.items():
            if month_name in text:
                year = today.year
                last_day = monthrange(year, month_number)[1]
                start_date = date(year, month_number, 1)
                end_date = date(year, month_number, last_day)
                break

    type_hint_words = ["bateria", "llanta", "pinchazo", "neumatico", "motor", "aceite", "choque", "accidente"]
    if any(word in text for word in type_hint_words) and not incident_type_value:
        warnings.append("No se reconocio el tipo de emergencia solicitado.")

    if payment_value == "card":
        warnings.append("El metodo de pago 'card' no esta disponible actualmente en reportes filtrados.")
        payment_value = None

    if not raw_text.strip():
        action = None
        warnings.append("No se recibio texto para interpretar.")

    def _extract_id(patterns: list[str]) -> Optional[int]:
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                try:
                    return int(match.group(1))
                except (TypeError, ValueError):
                    continue
        return None

    client_id = _extract_id(
        [
            r"\bcliente(?:\s+id)?\s+(\d+)\b",
            r"\bid\s+cliente\s+(\d+)\b",
        ]
    )
    workshop_id = _extract_id(
        [
            r"\btaller(?:\s+id)?\s+(\d+)\b",
            r"\bid\s+taller\s+(\d+)\b",
        ]
    )
    technician_id = _extract_id(
        [
            r"\btecnico(?:\s+id)?\s+(\d+)\b",
            r"\bid\s+tecnico\s+(\d+)\b",
        ]
    )
    vehicle_id = _extract_id(
        [
            r"\bvehiculo(?:\s+id)?\s+(\d+)\b",
            r"\bid\s+vehiculo\s+(\d+)\b",
        ]
    )

    filters = OperationalReportRequest(
        start_date=start_date,
        end_date=end_date,
        incident_type=incident_type_value,
        status=status_value,
        payment_method=payment_value,
        workshop_id=workshop_id,
        technician_id=technician_id,
        client_id=client_id,
        vehicle_id=vehicle_id,
    )
    return filters, action, warnings


def _to_datetime_range_start(value: Optional[date | datetime]) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    return datetime.combine(value, time.min)


def _to_datetime_range_end(value: Optional[date | datetime]) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    return datetime.combine(value, time.max)


def _parse_incident_status(value: Optional[str]) -> Optional[IncidentStatus]:
    if not value:
        return None
    normalized = value.strip().lower()
    for candidate in IncidentStatus:
        if candidate.value == normalized:
            return candidate
    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail="status invalido para el reporte operacional",
    )


def _parse_payment_method(value: Optional[str]) -> Optional[PaymentMethod]:
    if not value:
        return None
    normalized = value.strip().lower()
    for candidate in PaymentMethod:
        if candidate.value == normalized:
            return candidate
    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail="payment_method invalido para el reporte operacional",
    )


def _build_operational_report(
    payload: OperationalReportRequest,
    db: Session,
    current_user: User,
) -> OperationalReportResponse:
    allowed_roles = {UserRole.ADMIN, UserRole.WORKSHOP, UserRole.CLIENT}
    if current_user.role not in allowed_roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Tu rol no tiene acceso a reportes operacionales",
        )

    query = db.query(Incident).options(
        joinedload(Incident.user),
        joinedload(Incident.vehicle),
        joinedload(Incident.workshop),
        joinedload(Incident.technician),
        joinedload(Incident.payment),
    )

    role_scope = current_user.role.value
    workshop: Optional[Workshop] = None

    if current_user.role == UserRole.WORKSHOP:
        workshop = db.query(Workshop).filter(Workshop.owner_id == current_user.id).first()
        if not workshop:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No se encontro un taller asociado al usuario",
            )
        query = query.filter(Incident.workshop_id == workshop.id)
    elif current_user.role == UserRole.CLIENT:
        query = query.filter(Incident.user_id == current_user.id)

    start_dt = _to_datetime_range_start(payload.start_date)
    end_dt = _to_datetime_range_end(payload.end_date)
    if start_dt:
        query = query.filter(Incident.created_at >= start_dt)
    if end_dt:
        query = query.filter(Incident.created_at <= end_dt)

    if payload.incident_type:
        query = query.filter(Incident.classification == payload.incident_type.strip())

    parsed_status = _parse_incident_status(payload.status)
    if parsed_status:
        query = query.filter(Incident.status == parsed_status)

    if payload.technician_id is not None:
        query = query.filter(Incident.technician_id == payload.technician_id)

    if payload.payment_method:
        parsed_payment_method = _parse_payment_method(payload.payment_method)
        query = query.join(Payment, Incident.id == Payment.incident_id).filter(
            Payment.payment_method == parsed_payment_method
        )

    if current_user.role == UserRole.ADMIN:
        if payload.workshop_id is not None:
            query = query.filter(Incident.workshop_id == payload.workshop_id)
        if payload.client_id is not None:
            query = query.filter(Incident.user_id == payload.client_id)
        if payload.vehicle_id is not None:
            query = query.filter(Incident.vehicle_id == payload.vehicle_id)
    elif current_user.role == UserRole.WORKSHOP:
        if payload.technician_id is not None:
            technician = db.query(Technician).filter(
                Technician.id == payload.technician_id,
                Technician.workshop_id == workshop.id,
            ).first()
            if not technician:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="El tecnico no pertenece a tu taller",
                )
    elif current_user.role == UserRole.CLIENT and payload.vehicle_id is not None:
        query = query.filter(Incident.vehicle_id == payload.vehicle_id)

    incidents = query.order_by(Incident.created_at.desc(), Incident.id.desc()).all()

    counts = {st.value: 0 for st in IncidentStatus}
    total_amount = 0.0
    total_workshop_earnings = 0.0
    total_paid = 0
    total_unpaid = 0
    items: list[OperationalReportItem] = []

    for incident in incidents:
        counts[incident.status.value] += 1

        payment = incident.payment
        if payment:
            total_amount += float(payment.amount)
            total_workshop_earnings += float(payment.workshop_earnings)
            if payment.is_paid:
                total_paid += 1
            else:
                total_unpaid += 1

        items.append(
            OperationalReportItem(
                incident_id=incident.id,
                created_at=incident.created_at,
                updated_at=incident.updated_at,
                completed_at=incident.completed_at,
                status=incident.status,
                priority=incident.priority,
                classification=incident.classification or "Sin clasificar",
                description=incident.description,
                location_text=incident.location_text,
                client_id=incident.user_id,
                client_name=incident.user.full_name if incident.user else None,
                client_email=incident.user.email if incident.user else None,
                vehicle_id=incident.vehicle_id,
                vehicle_brand=incident.vehicle.brand if incident.vehicle else None,
                vehicle_model=incident.vehicle.model if incident.vehicle else None,
                vehicle_plate=incident.vehicle.plate if incident.vehicle else None,
                workshop_id=incident.workshop_id,
                workshop_name=incident.workshop.name if incident.workshop else None,
                technician_id=incident.technician_id,
                technician_name=incident.technician.name if incident.technician else None,
                payment_id=payment.id if payment else None,
                payment_amount=float(payment.amount) if payment else None,
                payment_method=payment.payment_method if payment else None,
                payment_is_paid=payment.is_paid if payment else None,
                commission_amount=float(payment.commission_amount) if payment else None,
                workshop_earnings=float(payment.workshop_earnings) if payment else None,
            )
        )

    summary = OperationalReportSummary(
        total_incidents=len(incidents),
        pending=counts[IncidentStatus.PENDING.value],
        waiting_offers=counts[IncidentStatus.WAITING_OFFERS.value],
        assigned=counts[IncidentStatus.ASSIGNED.value],
        accepted=counts[IncidentStatus.ACCEPTED.value],
        in_progress=counts[IncidentStatus.IN_PROGRESS.value],
        completed=counts[IncidentStatus.COMPLETED.value],
        cancelled=counts[IncidentStatus.CANCELLED.value],
        total_amount=total_amount,
        total_workshop_earnings=total_workshop_earnings,
        total_paid=total_paid,
        total_unpaid=total_unpaid,
    )

    applied_filters = OperationalReportAppliedFilters(
        start_date=start_dt.isoformat() if start_dt else None,
        end_date=end_dt.isoformat() if end_dt else None,
        workshop_id=payload.workshop_id if current_user.role == UserRole.ADMIN else (workshop.id if workshop else None),
        incident_type=payload.incident_type.strip() if payload.incident_type else None,
        status=parsed_status.value if parsed_status else None,
        technician_id=payload.technician_id,
        client_id=payload.client_id if current_user.role == UserRole.ADMIN else None,
        vehicle_id=payload.vehicle_id if current_user.role in [UserRole.ADMIN, UserRole.CLIENT] else None,
        payment_method=payload.payment_method.strip().lower() if payload.payment_method else None,
    )

    return OperationalReportResponse(
        role_scope=role_scope,
        applied_filters=applied_filters,
        summary=summary,
        items=items,
    )


def _generate_operational_pdf(report: OperationalReportResponse) -> bytes:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 40

    def new_page() -> None:
        nonlocal y
        pdf.showPage()
        y = height - 40

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(40, y, "Reporte Operacional")
    y -= 18

    pdf.setFont("Helvetica", 10)
    pdf.drawString(40, y, f"Rol: {report.role_scope}")
    y -= 14
    pdf.drawString(40, y, f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    y -= 20

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(40, y, "Filtros aplicados")
    y -= 14

    pdf.setFont("Helvetica", 9)
    filters_text = (
        f"start_date={report.applied_filters.start_date or 'N/A'} | "
        f"end_date={report.applied_filters.end_date or 'N/A'} | "
        f"workshop_id={report.applied_filters.workshop_id or 'N/A'} | "
        f"incident_type={report.applied_filters.incident_type or 'N/A'} | "
        f"status={report.applied_filters.status or 'N/A'} | "
        f"technician_id={report.applied_filters.technician_id or 'N/A'} | "
        f"client_id={report.applied_filters.client_id or 'N/A'} | "
        f"vehicle_id={report.applied_filters.vehicle_id or 'N/A'} | "
        f"payment_method={report.applied_filters.payment_method or 'N/A'}"
    )
    for line in simpleSplit(filters_text, "Helvetica", 9, width - 80):
        if y < 80:
            new_page()
        pdf.drawString(40, y, line)
        y -= 11

    y -= 8
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(40, y, "Resumen KPI")
    y -= 14

    pdf.setFont("Helvetica", 9)
    summary_lines = [
        f"total_incidents: {report.summary.total_incidents}",
        f"pending: {report.summary.pending}",
        f"waiting_offers: {report.summary.waiting_offers}",
        f"assigned: {report.summary.assigned}",
        f"accepted: {report.summary.accepted}",
        f"in_progress: {report.summary.in_progress}",
        f"completed: {report.summary.completed}",
        f"cancelled: {report.summary.cancelled}",
        f"total_amount: {report.summary.total_amount:.2f}",
        f"total_workshop_earnings: {report.summary.total_workshop_earnings:.2f}",
        f"total_paid: {report.summary.total_paid}",
        f"total_unpaid: {report.summary.total_unpaid}",
    ]
    for line in summary_lines:
        if y < 80:
            new_page()
        pdf.drawString(40, y, line)
        y -= 11

    y -= 8
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(40, y, "Detalle")
    y -= 14

    if not report.items:
        pdf.setFont("Helvetica", 10)
        pdf.drawString(40, y, "Sin registros para los filtros seleccionados.")
    else:
        for item in report.items:
            if y < 90:
                new_page()

            pdf.setFont("Helvetica-Bold", 9)
            pdf.drawString(40, y, f"Incidente #{item.incident_id} | {item.status.value} | {item.created_at.strftime('%Y-%m-%d %H:%M')}")
            y -= 11

            pdf.setFont("Helvetica", 8)
            detail_line = (
                f"Cliente: {item.client_name or '-'} | Taller: {item.workshop_name or '-'} | "
                f"Vehiculo: {item.vehicle_brand or '-'} {item.vehicle_model or '-'} ({item.vehicle_plate or '-'}) | "
                f"Monto: {(item.payment_amount or 0):.2f} | "
                f"Pago: {item.payment_method.value if item.payment_method else '-'} | Pagado: {item.payment_is_paid}"
            )
            for line in simpleSplit(detail_line, "Helvetica", 8, width - 80):
                if y < 80:
                    new_page()
                pdf.drawString(40, y, line)
                y -= 10

            for line in simpleSplit(f"Descripcion: {item.description}", "Helvetica", 8, width - 80):
                if y < 80:
                    new_page()
                pdf.drawString(40, y, line)
                y -= 10

            y -= 6

    pdf.save()
    data = buffer.getvalue()
    buffer.close()
    return data


def _generate_operational_excel(report: OperationalReportResponse) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    ws_detail = wb.create_sheet("Detalle")

    ws_summary["A1"] = "role_scope"
    ws_summary["B1"] = report.role_scope

    summary_rows = [
        ("start_date", report.applied_filters.start_date),
        ("end_date", report.applied_filters.end_date),
        ("workshop_id", report.applied_filters.workshop_id),
        ("incident_type", report.applied_filters.incident_type),
        ("status", report.applied_filters.status),
        ("technician_id", report.applied_filters.technician_id),
        ("client_id", report.applied_filters.client_id),
        ("vehicle_id", report.applied_filters.vehicle_id),
        ("payment_method", report.applied_filters.payment_method),
        ("total_incidents", report.summary.total_incidents),
        ("pending", report.summary.pending),
        ("waiting_offers", report.summary.waiting_offers),
        ("assigned", report.summary.assigned),
        ("accepted", report.summary.accepted),
        ("in_progress", report.summary.in_progress),
        ("completed", report.summary.completed),
        ("cancelled", report.summary.cancelled),
        ("total_amount", report.summary.total_amount),
        ("total_workshop_earnings", report.summary.total_workshop_earnings),
        ("total_paid", report.summary.total_paid),
        ("total_unpaid", report.summary.total_unpaid),
    ]
    for idx, (key, value) in enumerate(summary_rows, start=2):
        ws_summary.cell(row=idx, column=1, value=key)
        ws_summary.cell(row=idx, column=2, value=value)

    headers = [
        "incident_id",
        "created_at",
        "status",
        "classification",
        "description",
        "location_text",
        "client_name",
        "client_email",
        "vehicle_brand",
        "vehicle_model",
        "vehicle_plate",
        "workshop_name",
        "technician_name",
        "payment_amount",
        "payment_method",
        "payment_is_paid",
        "commission_amount",
        "workshop_earnings",
    ]
    for col, header in enumerate(headers, start=1):
        ws_detail.cell(row=1, column=col, value=header)

    for row, item in enumerate(report.items, start=2):
        ws_detail.cell(row=row, column=1, value=item.incident_id)
        ws_detail.cell(row=row, column=2, value=item.created_at.isoformat())
        ws_detail.cell(row=row, column=3, value=item.status.value)
        ws_detail.cell(row=row, column=4, value=item.classification)
        ws_detail.cell(row=row, column=5, value=item.description)
        ws_detail.cell(row=row, column=6, value=item.location_text)
        ws_detail.cell(row=row, column=7, value=item.client_name)
        ws_detail.cell(row=row, column=8, value=item.client_email)
        ws_detail.cell(row=row, column=9, value=item.vehicle_brand)
        ws_detail.cell(row=row, column=10, value=item.vehicle_model)
        ws_detail.cell(row=row, column=11, value=item.vehicle_plate)
        ws_detail.cell(row=row, column=12, value=item.workshop_name)
        ws_detail.cell(row=row, column=13, value=item.technician_name)
        ws_detail.cell(row=row, column=14, value=item.payment_amount)
        ws_detail.cell(row=row, column=15, value=item.payment_method.value if item.payment_method else None)
        ws_detail.cell(row=row, column=16, value=item.payment_is_paid)
        ws_detail.cell(row=row, column=17, value=item.commission_amount)
        ws_detail.cell(row=row, column=18, value=item.workshop_earnings)

    output = BytesIO()
    wb.save(output)
    data = output.getvalue()
    output.close()
    return data


@router.post("/operational/query", response_model=OperationalReportResponse)
def query_operational_report(
    payload: OperationalReportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return _build_operational_report(payload=payload, db=db, current_user=current_user)


@router.post("/operational/export/pdf")
def export_operational_report_pdf(
    payload: OperationalReportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    report = _build_operational_report(payload=payload, db=db, current_user=current_user)
    pdf_bytes = _generate_operational_pdf(report)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="reporte_operacional.pdf"'},
    )


@router.post("/operational/export/excel")
def export_operational_report_excel(
    payload: OperationalReportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    report = _build_operational_report(payload=payload, db=db, current_user=current_user)
    excel_bytes = _generate_operational_excel(report)
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="reporte_operacional.xlsx"'},
    )


@router.post("/operational/voice-parse", response_model=VoiceReportParseResponse)
def parse_operational_voice_command(
    payload: VoiceReportParseRequest,
    current_user: User = Depends(get_current_user),
):
    allowed_roles = {UserRole.ADMIN, UserRole.WORKSHOP, UserRole.CLIENT}
    if current_user.role not in allowed_roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Tu rol no tiene acceso a comandos de voz para reportes",
        )

    filters, action, warnings = _parse_voice_filters(payload.text or "")

    # Seguridad por rol: voz no debe proponer filtros de alcance prohibido
    if current_user.role == UserRole.WORKSHOP:
        if filters.client_id is not None:
            warnings.append("El filtro client_id no está permitido para el rol workshop.")
            filters.client_id = None
        if filters.workshop_id is not None:
            warnings.append("El filtro workshop_id no está permitido para el rol workshop.")
            filters.workshop_id = None
        if filters.vehicle_id is not None:
            warnings.append("El filtro vehicle_id no está permitido para el rol workshop.")
            filters.vehicle_id = None
    elif current_user.role == UserRole.CLIENT:
        if filters.client_id is not None:
            warnings.append("El filtro client_id no está permitido para el rol client.")
            filters.client_id = None
        if filters.workshop_id is not None:
            warnings.append("El filtro workshop_id no está permitido para el rol client.")
            filters.workshop_id = None
        if filters.technician_id is not None:
            warnings.append("El filtro technician_id no está permitido para el rol client.")
            filters.technician_id = None

    return VoiceReportParseResponse(
        recognized_text=payload.text,
        filters=filters,
        action=action,
        warnings=warnings,
    )








